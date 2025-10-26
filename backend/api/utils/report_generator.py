"""Report generation utilities for PDF and Excel exports."""
import io
import calendar
from datetime import datetime
from typing import Dict, List, Any
import urllib.request
import tempfile
import os


def format_month_period(period_start: datetime, period_end: datetime) -> str:
    """
    Format period as '1st Month - Last_Day Month Year'.
    Example: '1st August - 31st August 2025'
    """
    month_name = period_start.strftime("%B")
    year = period_start.strftime("%Y")
    last_day = period_end.day

    # Add ordinal suffix
    def ordinal(n):
        suffix = ['th', 'st', 'nd', 'rd', 'th'][min(n % 10, 4)]
        if 11 <= (n % 100) <= 13:
            suffix = 'th'
        return f"{n}{suffix}"

    return f"{ordinal(1)} {month_name} - {ordinal(last_day)} {month_name} {year}"


def get_month_name(period_start: datetime) -> str:
    """Get month name from datetime."""
    return period_start.strftime("%B %Y")


def calculate_totals(partner_allocations: List[Dict]) -> Dict[str, float]:
    """Calculate totals for all numeric columns."""
    totals = {
        "gross_volume": 0,
        "net_volume_standard": 0,
        "allocated_volume": 0,
        "ownership_percent": 0,
    }

    for allocation in partner_allocations:
        totals["gross_volume"] += allocation.get("gross_volume", 0)
        totals["net_volume_standard"] += allocation.get("net_volume_standard", 0)
        totals["allocated_volume"] += allocation.get("allocated_volume", 0)
        totals["ownership_percent"] += allocation.get("ownership_percent", 0)

    return totals


def download_logo(logo_url: str) -> str:
    """Download logo from URL and return temporary file path."""
    if not logo_url:
        return None

    try:
        # Create temporary file
        temp_dir = tempfile.gettempdir()
        logo_path = os.path.join(temp_dir, 'flowshare_logo.webp')

        # Download logo
        urllib.request.urlretrieve(logo_url, logo_path)
        return logo_path
    except Exception as e:
        print(f"Failed to download logo: {e}")
        return None


def generate_pdf_report(reconciliation_id: str, reconciliation_data: Dict, result: Dict, logo_url: str = None) -> bytes:
    """Generate professional PDF report matching sample design."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.platypus.flowables import HRFlowable
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
        from reportlab.pdfgen import canvas
    except ImportError:
        raise ImportError("reportlab package is required for PDF generation")

    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        topMargin=0.75*inch,
        bottomMargin=0.75*inch,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch
    )

    # Container for the 'Flowable' objects
    elements = []

    # Define styles
    styles = getSampleStyleSheet()

    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.white,
        spaceAfter=6,
        alignment=TA_LEFT,
        fontName='Helvetica-Bold'
    )

    # Period style
    period_style = ParagraphStyle(
        'PeriodStyle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.white,
        spaceAfter=20,
        alignment=TA_LEFT,
        fontName='Helvetica'
    )

    # Heading style
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        spaceBefore=20,
        fontName='Helvetica-Bold'
    )

    # Body text style
    body_style = ParagraphStyle(
        'BodyStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#374151'),
        alignment=TA_JUSTIFY,
        spaceAfter=8,
        leading=14
    )

    # Get period formatted
    period_start = reconciliation_data.get("period_start")
    period_end = reconciliation_data.get("period_end")
    if isinstance(period_start, str):
        period_start = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
    if isinstance(period_end, str):
        period_end = datetime.fromisoformat(period_end.replace('Z', '+00:00'))

    period_text = format_month_period(period_start, period_end)
    month_name = get_month_name(period_start)

    # Download logo if URL provided
    logo_path = download_logo(logo_url) if logo_url else None

    # Header section with blue background and logo
    if logo_path:
        try:
            # Create logo image (small, for top-left corner)
            logo_img = Image(logo_path, width=0.6*inch, height=0.6*inch)
            logo_img.hAlign = 'LEFT'

            # Header with logo and text
            header_table_data = [
                [logo_img, Paragraph("Reconciliation Report", title_style)],
                ['', Paragraph(period_text, period_style)]
            ]

            header_table = Table(header_table_data, colWidths=[0.8*inch, 6.2*inch])
            header_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#4169E1')),  # Royal blue
                ('LEFTPADDING', (0, 0), (-1, -1), 20),
                ('RIGHTPADDING', (0, 0), (-1, -1), 30),
                ('TOPPADDING', (0, 0), (-1, -1), 25),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 25),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('VALIGN', (0, 0), (0, 0), 'TOP'),  # Logo at top
            ]))
        except Exception as e:
            print(f"Failed to add logo to PDF: {e}")
            # Fallback without logo
            header_table_data = [
                [Paragraph("Reconciliation Report", title_style)],
                [Paragraph(period_text, period_style)]
            ]
            header_table = Table(header_table_data, colWidths=[7*inch])
            header_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#4169E1')),
                ('LEFTPADDING', (0, 0), (-1, -1), 30),
                ('RIGHTPADDING', (0, 0), (-1, -1), 30),
                ('TOPPADDING', (0, 0), (-1, -1), 30),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 30),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
    else:
        # Header without logo
        header_table_data = [
            [Paragraph("Reconciliation Report", title_style)],
            [Paragraph(period_text, period_style)]
        ]
        header_table = Table(header_table_data, colWidths=[7*inch])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#4169E1')),
            ('LEFTPADDING', (0, 0), (-1, -1), 30),
            ('RIGHTPADDING', (0, 0), (-1, -1), 30),
            ('TOPPADDING', (0, 0), (-1, -1), 30),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 30),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

    elements.append(header_table)
    elements.append(Spacer(1, 0.4*inch))

    # Calculate totals and metrics
    partner_allocations = result.get("partner_allocations", [])
    totals = calculate_totals(partner_allocations)
    terminal_volume = reconciliation_data.get("terminal_volume", 0)

    # Calculate gain/loss
    total_input = totals["gross_volume"]
    gain_loss = total_input - terminal_volume
    shrinkage_percent = result.get("shrinkage_percent", 0)
    allocated_volume = totals["allocated_volume"]

    # Key metrics cards (2x3 grid)
    metrics_data = [
        [
            ["TOTAL PARTNERS", f"{len(partner_allocations)}"],
            ["INPUT VOLUME", f"{total_input:,.0f} BBL"],
            ["TERMINAL VOLUME", f"{terminal_volume:,.0f} BBL"]
        ],
        [
            ["VOLUME LOSS/GAIN", f"{gain_loss:,.2f} BBL"],
            ["SHRINKAGE", f"{shrinkage_percent:.2f}%"],
            ["ALLOCATED VOLUME", f"{allocated_volume:,.2f} BBL"]
        ]
    ]

    # Create metric cards
    for row in metrics_data:
        row_tables = []
        for metric in row:
            metric_table = Table(
                [[Paragraph(metric[0], ParagraphStyle('MetricLabel', fontSize=9, textColor=colors.HexColor('#6B7280')))],
                 [Paragraph(metric[1], ParagraphStyle('MetricValue', fontSize=14, textColor=colors.black, fontName='Helvetica-Bold'))]],
                colWidths=[2.2*inch]
            )
            metric_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F9FAFB')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            row_tables.append(metric_table)

        # Arrange in row
        row_table = Table([row_tables], colWidths=[2.3*inch, 2.3*inch, 2.3*inch], spaceBefore=8, spaceAfter=8)
        row_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(row_table)

    elements.append(Spacer(1, 0.3*inch))

    # Executive Summary
    elements.append(Paragraph(f"Executive Summary: Joint Venture Reconciliation - {month_name}", heading_style))

    summary_text = f"""This report summarizes the reconciliation of crude oil volumes for the {month_name} period.
    Total input volume was {total_input:,.2f} BBL, with a reconciled terminal volume of {terminal_volume:,.2f} BBL.

    Key findings indicate a total shrinkage factor of {shrinkage_percent:.2f}%. The allocated volume totals {allocated_volume:,.2f} BBL
    across {len(partner_allocations)} partners. The allocation was performed using the {result.get('allocation_model_used', 'API MPMS 11.1')}
    methodology, ensuring accurate and fair distribution based on each partner's contribution."""

    elements.append(Paragraph(summary_text, body_style))
    elements.append(Spacer(1, 0.3*inch))

    # Partner Allocations Table
    allocation_table_data = [
        [
            Paragraph("<b>PARTNER</b>", ParagraphStyle('TableHeader', fontSize=9, textColor=colors.HexColor('#374151'), fontName='Helvetica-Bold')),
            Paragraph("<b>INPUT VOLUME</b>", ParagraphStyle('TableHeader', fontSize=9, textColor=colors.HexColor('#374151'), fontName='Helvetica-Bold', alignment=TA_CENTER)),
            Paragraph("<b>ALLOCATED VOLUME</b>", ParagraphStyle('TableHeader', fontSize=9, textColor=colors.HexColor('#374151'), fontName='Helvetica-Bold', alignment=TA_CENTER)),
            Paragraph("<b>GAIN/LOSS</b>", ParagraphStyle('TableHeader', fontSize=9, textColor=colors.HexColor('#374151'), fontName='Helvetica-Bold', alignment=TA_CENTER)),
            Paragraph("<b>SHARE</b>", ParagraphStyle('TableHeader', fontSize=9, textColor=colors.HexColor('#374151'), fontName='Helvetica-Bold', alignment=TA_CENTER)),
            Paragraph("<b>EFFICIENCY</b>", ParagraphStyle('TableHeader', fontSize=9, textColor=colors.HexColor('#374151'), fontName='Helvetica-Bold', alignment=TA_CENTER))
        ]
    ]

    for allocation in partner_allocations:
        partner_gross = allocation.get("gross_volume", 0)
        partner_allocated = allocation.get("allocated_volume", 0)
        partner_gain_loss = partner_gross - partner_allocated
        partner_share = allocation.get("ownership_percent", 0)
        partner_efficiency = (partner_allocated / partner_gross * 100) if partner_gross > 0 else 0

        allocation_table_data.append([
            allocation.get("partner_name", "N/A"),
            f"{partner_gross:,.0f} bbl",
            f"{partner_allocated:,.2f} bbl",
            f"{partner_gain_loss:,.2f} bbl",
            f"{partner_share:.2f}%",
            f"{partner_efficiency:.1f}%"
        ])

    # Add TOTAL row
    total_efficiency = (allocated_volume / total_input * 100) if total_input > 0 else 0
    allocation_table_data.append([
        Paragraph("<b>TOTAL</b>", ParagraphStyle('TotalLabel', fontSize=10, fontName='Helvetica-Bold')),
        f"{total_input:,.0f} bbl",
        f"{allocated_volume:,.2f} bbl",
        f"{gain_loss:,.2f} bbl",
        f"{totals['ownership_percent']:.2f}%",
        "—"
    ])

    allocation_table = Table(allocation_table_data, colWidths=[1.4*inch, 1.15*inch, 1.3*inch, 1.15*inch, 0.9*inch, 1.0*inch])
    allocation_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),

        # Data rows alternating background
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAFAFA')]),

        # Total row styling
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#D1D5DB')),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
    ]))
    elements.append(allocation_table)

    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_text = f"Generated on {datetime.now().strftime('%m/%d/%Y, %I:%M:%S %p')} • FlowShare Reconciliation System<br/>Run Date: {datetime.now().strftime('%m/%d/%Y')}"
    elements.append(Paragraph(footer_text, ParagraphStyle('Footer', fontSize=8, textColor=colors.HexColor('#9CA3AF'), alignment=TA_CENTER)))

    # Build PDF
    doc.build(elements)

    # Get PDF content
    pdf_content = buffer.getvalue()
    buffer.close()

    return pdf_content


def generate_excel_report(reconciliation_id: str, reconciliation_data: Dict, result: Dict) -> bytes:
    """Generate Excel report with 3 sheets: Summary, Allocations, Calculations."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl package is required for Excel generation")

    wb = Workbook()

    # Get period info
    period_start = reconciliation_data.get("period_start")
    period_end = reconciliation_data.get("period_end")
    if isinstance(period_start, str):
        period_start = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
    if isinstance(period_end, str):
        period_end = datetime.fromisoformat(period_end.replace('Z', '+00:00'))

    month_name = get_month_name(period_start)
    period_text = format_month_period(period_start, period_end)

    # Calculate totals
    partner_allocations = result.get("partner_allocations", [])
    totals = calculate_totals(partner_allocations)
    terminal_volume = reconciliation_data.get("terminal_volume", 0)

    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4169E1', end_color='4169E1', fill_type='solid')
    title_font = Font(name='Calibri', size=14, bold=True)
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"

    ws1['A1'] = 'FlowShare V2 - Reconciliation Report'
    ws1['A1'].font = title_font
    ws1['A2'] = f'Period: {period_text}'
    ws1['A2'].font = bold_font
    ws1['A3'] = f'Month: {month_name}'
    ws1['A3'].font = bold_font

    ws1['A5'] = 'Reconciliation ID:'
    ws1['B5'] = reconciliation_id
    ws1['A6'] = 'Status:'
    ws1['B6'] = reconciliation_data.get("status", "").upper()
    ws1['A7'] = 'Allocation Model:'
    ws1['B7'] = result.get("allocation_model_used", "N/A")

    ws1['A9'] = 'Key Metrics'
    ws1['A9'].font = bold_font

    metrics = [
        ['Total Partners', len(partner_allocations)],
        ['Input Volume (BBL)', totals["gross_volume"]],
        ['Terminal Volume (BBL)', terminal_volume],
        ['Allocated Volume (BBL)', totals["allocated_volume"]],
        ['Volume Loss/Gain (BBL)', totals["gross_volume"] - terminal_volume],
        ['Shrinkage (%)', result.get("shrinkage_percent", 0)],
        ['Total Ownership (%)', totals["ownership_percent"]]
    ]

    for i, (metric, value) in enumerate(metrics, start=10):
        ws1[f'A{i}'] = metric
        ws1[f'B{i}'] = value
        ws1[f'A{i}'].font = bold_font
        if isinstance(value, (int, float)):
            ws1[f'B{i}'].number_format = '#,##0.00'

    # Sheet 2: Partner Allocations
    ws2 = wb.create_sheet("Partner Allocations")

    ws2['A1'] = f'Partner Allocations - {month_name}'
    ws2['A1'].font = title_font

    headers = ['Partner Name', 'Input Volume (bbl)', 'Allocated Volume (bbl)', 'Gain/Loss (bbl)', 'Ownership Share (%)', 'Efficiency (%)']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row = 4
    for allocation in partner_allocations:
        partner_gross = allocation.get("gross_volume", 0)
        partner_allocated = allocation.get("allocated_volume", 0)
        partner_gain_loss = partner_gross - partner_allocated
        partner_share = allocation.get("ownership_percent", 0)
        partner_efficiency = (partner_allocated / partner_gross * 100) if partner_gross > 0 else 0

        ws2.cell(row=row, column=1, value=allocation.get("partner_name", "N/A")).border = border
        ws2.cell(row=row, column=2, value=partner_gross).border = border
        ws2.cell(row=row, column=2).number_format = '#,##0.00'
        ws2.cell(row=row, column=3, value=partner_allocated).border = border
        ws2.cell(row=row, column=3).number_format = '#,##0.00'
        ws2.cell(row=row, column=4, value=partner_gain_loss).border = border
        ws2.cell(row=row, column=4).number_format = '#,##0.00'
        ws2.cell(row=row, column=5, value=partner_share).border = border
        ws2.cell(row=row, column=5).number_format = '0.00'
        ws2.cell(row=row, column=6, value=partner_efficiency).border = border
        ws2.cell(row=row, column=6).number_format = '0.0'
        row += 1

    # Add total row
    total_row = row
    ws2.cell(row=total_row, column=1, value='TOTAL').font = bold_font
    ws2.cell(row=total_row, column=1).border = border
    ws2.cell(row=total_row, column=2, value=totals["gross_volume"]).font = bold_font
    ws2.cell(row=total_row, column=2).number_format = '#,##0.00'
    ws2.cell(row=total_row, column=2).border = border
    ws2.cell(row=total_row, column=3, value=totals["allocated_volume"]).font = bold_font
    ws2.cell(row=total_row, column=3).number_format = '#,##0.00'
    ws2.cell(row=total_row, column=3).border = border
    ws2.cell(row=total_row, column=4, value=totals["gross_volume"] - terminal_volume).font = bold_font
    ws2.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws2.cell(row=total_row, column=4).border = border
    ws2.cell(row=total_row, column=5, value=totals["ownership_percent"]).font = bold_font
    ws2.cell(row=total_row, column=5).number_format = '0.00'
    ws2.cell(row=total_row, column=5).border = border
    ws2.cell(row=total_row, column=6, value='—').border = border

    # Adjust column widths
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 3: Intermediate Calculations
    ws3 = wb.create_sheet("Intermediate Calculations")

    ws3['A1'] = f'Detailed Calculations - {month_name}'
    ws3['A1'].font = title_font

    calc_headers = ['Partner Name', 'Gross Volume (bbl)', 'BSW %', 'Water Cut Factor',
                    'Net Volume Observed (bbl)', 'Temp Correction Factor', 'API Correction Factor',
                    'Net Volume Standard (bbl)', 'Ownership %', 'Allocated Volume (bbl)']

    for col, header in enumerate(calc_headers, start=1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    row = 4
    for allocation in partner_allocations:
        ws3.cell(row=row, column=1, value=allocation.get("partner_name", "N/A")).border = border
        ws3.cell(row=row, column=2, value=allocation.get("gross_volume", 0)).border = border
        ws3.cell(row=row, column=2).number_format = '#,##0.00'
        ws3.cell(row=row, column=3, value=allocation.get("bsw_percent", 0)).border = border
        ws3.cell(row=row, column=3).number_format = '0.0000'
        ws3.cell(row=row, column=4, value=allocation.get("water_cut_factor", 0)).border = border
        ws3.cell(row=row, column=4).number_format = '0.000000'
        ws3.cell(row=row, column=5, value=allocation.get("net_volume_observed", 0)).border = border
        ws3.cell(row=row, column=5).number_format = '#,##0.00'
        ws3.cell(row=row, column=6, value=allocation.get("temperature_correction_factor", 0)).border = border
        ws3.cell(row=row, column=6).number_format = '0.000000'
        ws3.cell(row=row, column=7, value=allocation.get("api_correction_factor", 0)).border = border
        ws3.cell(row=row, column=7).number_format = '0.000000'
        ws3.cell(row=row, column=8, value=allocation.get("net_volume_standard", 0)).border = border
        ws3.cell(row=row, column=8).number_format = '#,##0.00'
        ws3.cell(row=row, column=9, value=allocation.get("ownership_percent", 0)).border = border
        ws3.cell(row=row, column=9).number_format = '0.0000'
        ws3.cell(row=row, column=10, value=allocation.get("allocated_volume", 0)).border = border
        ws3.cell(row=row, column=10).number_format = '#,##0.00'
        row += 1

    # Adjust column widths
    for col in range(1, 11):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_content = excel_buffer.getvalue()
    excel_buffer.close()

    return excel_content
